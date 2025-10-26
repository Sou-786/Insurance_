# excel_mcp_server.py
from fastmcp import FastMCP
import openpyxl
import os

# Initialize MCP app
mcp = FastMCP("c")

# === CONFIGURATION ===
EXCEL_PATH = r"C:\Users\User\Downloads\Insurance.xlsx"

if not os.path.exists(EXCEL_PATH):
    raise FileNotFoundError(f"Excel file not found: {EXCEL_PATH}")


# === Helper Functions ===
def load_workbook(file_path: str):
    """Load Excel workbook using openpyxl."""
    return openpyxl.load_workbook(file_path, data_only=True)


def extract_tables_from_sheet(sheet):
    """
    Extract consecutive non-empty rows as separate tables.
    Returns a list of tables (each table is a list of rows).
    """
    tables = []
    current_table = []

    for row in sheet.iter_rows(values_only=True):
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            current_table.append([cell if cell is not None else "" for cell in row])
        elif current_table:
            tables.append(current_table)
            current_table = []

    if current_table:
        tables.append(current_table)

    return tables


# === MCP Tools ===

@mcp.tool(name="list_excel_sheets")
def list_sheets() -> dict:
    """
    Lists all sheets in the Excel file.
    """
    wb = load_workbook(EXCEL_PATH)
    return {"excel_file": EXCEL_PATH, "sheets": wb.sheetnames}


@mcp.tool(name="get_sheet_preview")
def get_sheet_preview(sheet_name: str, table_limit: int = 2, row_preview: int = 5) -> dict:
    """
    Extracts table-like sections from a specific sheet.
    Parameters:
        sheet_name: Name of the sheet to read
        table_limit: Number of tables to preview
        row_preview: Number of rows to show per table
    """
    wb = load_workbook(EXCEL_PATH)
    if sheet_name not in wb.sheetnames:
        return {"error": f"Sheet '{sheet_name}' not found."}

    sheet = wb[sheet_name]
    tables = extract_tables_from_sheet(sheet)
    preview = []

    for i, table in enumerate(tables[:table_limit]):
        preview.append({
            "table_index": i + 1,
            "rows": len(table),
            "cols": len(table[0]) if table else 0,
            "data": table[:row_preview]
        })

    return {
        "sheet_name": sheet_name,
        "detected_tables": len(tables),
        "preview": preview
    }


# === Run MCP ===
if __name__ == "__main__":
    mcp.run()
